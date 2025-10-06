import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.db import connection
from django.contrib import admin
from django.contrib.admin import RelatedOnlyFieldListFilter
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from .models import (
    Usuario, PerfilCoordinador, PerfilSoporte, 
    PerfilMonitor, PerfilOperador, Acta, Actividad, ObservacionAdicional, DetalleAdicional
)



# Cambiar el encabezado que aparece en la parte superior del admin
admin.site.site_header = "Sistema de Administración de Operadores"


# -------------------
# Admin para Usuario
# -------------------
class UsuarioAdmin(BaseUserAdmin):
    model = Usuario
    list_display = ('username', 'rol', 'email', 'is_staff', 'is_active')
    list_filter = ('rol', 'is_staff', 'is_active')
    search_fields = ('username', 'email')
    ordering = ('username',)

    fieldsets = (
        (None, {'fields': ('username', 'password', 'rol', 'email')}),
        ('Permisos', {'fields': ('is_staff', 'is_active', 'groups', 'user_permissions')}),
    )

    add_fieldsets = (
        (None, {
            'classes': ('wide',),
            'fields': ('username', 'rol', 'email', 'password1', 'password2', 'is_staff', 'is_active')}
        ),
    )

# -------------------
# Admin para perfiles
# -------------------
@admin.register(PerfilCoordinador)
class PerfilCoordinadorAdmin(admin.ModelAdmin):
    list_display = ('codigo', 'nombres', 'paterno', 'materno', 'celular', 'usuario')
    search_fields = ('nombres', 'paterno', 'materno', 'usuario__username')


@admin.register(PerfilMonitor)
class PerfilMonitorAdmin(admin.ModelAdmin):
    list_display = ('codigo', 'nombres', 'paterno', 'materno', 'celular', 'usuario')
    search_fields = ('nombres', 'paterno', 'materno', 'usuario__username')


@admin.register(PerfilSoporte)
class PerfilSoporteAdmin(admin.ModelAdmin):
    list_display = ('codigo', 'nombres', 'paterno', 'materno', 'celular', 'usuario')
    search_fields = ('nombres', 'paterno', 'materno', 'usuario__username')


# -------------------
# Admin para PerfilOperador
# -------------------
@admin.register(PerfilOperador)
class PerfilOperadorAdmin(admin.ModelAdmin):
    list_display = (
        'ci', 'usuario', 'nombre_completo', 'celular',
        'tipo_personal', 
        'get_coordinador', 'get_coordinador_celular', 
        'get_monitor', 'get_monitor_celular',
        'get_soporte', 'get_soporte_celular',
    
        # Actas
        'get_provincia', 'get_municipio', 'get_localidad', 'get_recinto',
        'get_mesa1', 'get_mesa2', 'get_mesa3', 'get_mesa4',
        'get_mesa5', 'get_mesa6', 'get_mesa7', 'get_mesa8',
        # Detalle adicional
        'detalle_usuario_apk', 'get_actas_asignadas', 'detalle_transmitido', 'detalle_no_transmitido',
        # Observaciones adicionales
        'observacion1',
    )
    search_fields = ('nombres', 'paterno', 'materno', 'ci', 'celular')
    list_filter = ('tipo_personal', 'coordinador', 'monitor', 'soporte')

    # -------------------
    # Nombre completo
    # -------------------
    def nombre_completo(self, obj):
        return f"{obj.nombres} {obj.paterno} {obj.materno}"
    nombre_completo.short_description = "Nombre Completo"

    # -------------------
    # Coordinador / Soporte / Monitor
    # -------------------
    def get_coordinador(self, obj):
        if obj.coordinador:
            c = obj.coordinador
            return f"{c.nombres} {c.paterno} {c.materno}"
        return "—"
    get_coordinador.short_description = "Coordinador"

    def get_soporte(self, obj):
        if obj.soporte:
            s = obj.soporte
            return f"{s.nombres} {s.paterno} {s.materno}"
        return "—"
    get_soporte.short_description = "Soporte"

    def get_monitor(self, obj):
        if obj.monitor:
            m = obj.monitor
            return f"{m.nombres} {m.paterno} {m.materno}"
        return "—"
    get_monitor.short_description = "Monitor"

    # -------------------
    # Celulares de relaciones
    # -------------------
    def get_coordinador_celular(self, obj):
        return obj.coordinador.celular if obj.coordinador else "—"
    get_coordinador_celular.short_description = "Celular Coordinador"

    def get_soporte_celular(self, obj):
        return obj.soporte.celular if obj.soporte else "—"
    get_soporte_celular.short_description = "Celular Soporte"

    def get_monitor_celular(self, obj):
        return obj.monitor.celular if obj.monitor else "—"
    get_monitor_celular.short_description = "Celular Monitor"

    # -------------------
    # Acta relacionada
    # -------------------
    def get_acta(self, obj):
        return getattr(obj, 'acta_operador', None)

    def get_actas_asignadas(self, obj):
        acta = self.get_acta(obj)
        return acta.actas_asignadas if acta else "-"
    get_actas_asignadas.short_description = "Actas Asignadas"

    def get_provincia(self, obj):
        acta = self.get_acta(obj)
        return acta.provincia if acta else "-"
    get_provincia.short_description = "Provincia"

    def get_municipio(self, obj):
        acta = self.get_acta(obj)
        return acta.municipio if acta else "-"
    get_municipio.short_description = "Municipio"

    def get_localidad(self, obj):
        acta = self.get_acta(obj)
        return acta.localidad if acta else "-"
    get_localidad.short_description = "Localidad"

    def get_recinto(self, obj):
        acta = self.get_acta(obj)
        return acta.recinto if acta else "-"
    get_recinto.short_description = "Recinto"

    def get_mesa1(self, obj):
        acta = self.get_acta(obj)
        return acta.mesa1 if acta else "-"
    get_mesa1.short_description = "Mesa 1"

    def get_mesa2(self, obj):
        acta = self.get_acta(obj)
        return acta.mesa2 if acta else "-"
    get_mesa2.short_description = "Mesa 2"

    def get_mesa3(self, obj):
        acta = self.get_acta(obj)
        return acta.mesa3 if acta else "-"
    get_mesa3.short_description = "Mesa 3"

    def get_mesa4(self, obj):
        acta = self.get_acta(obj)
        return acta.mesa4 if acta else "-"
    get_mesa4.short_description = "Mesa 4"

    def get_mesa5(self, obj):
        acta = self.get_acta(obj)
        return acta.mesa5 if acta else "-"
    get_mesa5.short_description = "Mesa 5"

    def get_mesa6(self, obj):
        acta = self.get_acta(obj)
        return acta.mesa6 if acta else "-"
    get_mesa6.short_description = "Mesa 6"

    def get_mesa7(self, obj):
        acta = self.get_acta(obj)
        return acta.mesa7 if acta else "-"
    get_mesa7.short_description = "Mesa 7"

    def get_mesa8(self, obj):
        acta = self.get_acta(obj)
        return acta.mesa8 if acta else "-"
    get_mesa8.short_description = "Mesa 8"

    # -------------------
    # Observaciones adicionales
    # -------------------
    def get_observacion(self, obj, index):
        obs = list(obj.observaciones_adicionales.all())
        if index < len(obs):
            return f"{obs[index].observacion}"
        return "-"
    def observacion1(self, obj): return self.get_observacion(obj, 0)
    observacion1.short_description = "Observación Adicional"

    # -------------------
    # 🔹 Detalle adicional
    # -------------------
    def get_detalle(self, obj):
        return getattr(obj, 'detalle_adicional', None)

    def detalle_usuario_apk(self, obj):
        detalle = self.get_detalle(obj)
        return detalle.usuario_apk if detalle else "-"
    detalle_usuario_apk.short_description = "Usuario APK"

    def detalle_transmitido(self, obj):
        detalle = self.get_detalle(obj)
        return detalle.transmitido if detalle else "-"
    detalle_transmitido.short_description = "Transmitido"

    def detalle_no_transmitido(self, obj):
        detalle = self.get_detalle(obj)
        return detalle.no_transmitido if detalle else "-"
    detalle_no_transmitido.short_description = "No Transmitido"

    # -------------------
    # Filtrado por operador
    # -------------------
    def get_queryset(self, request):
        qs = super().get_queryset(request)
        if hasattr(request.user, 'rol') and request.user.rol == 'operador':
            return qs.filter(usuario=request.user)
        return qs

    
    actions = ["exportar_excel", "consulta_personalizada"]
    # -------------------
    # Acción 1: Exportar Excel
    # -------------------
    def exportar_excel(self, request, queryset):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Perfil Operador"

        headers = [field.replace("_", " ").title() for field in self.list_display]
        ws.append(headers)

        for obj in queryset:
            row = []
            for field in self.list_display:
                attr = getattr(self, field, None)
                if callable(attr):
                    value = attr(obj)
                else:
                    value = getattr(obj, field, "")
                row.append(str(value))
            ws.append(row)

        for i, column_cells in enumerate(ws.columns, 1):
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(i)].width = length + 2

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="perfil_operadores.xlsx"'
        wb.save(response)
        return response

    exportar_excel.short_description = "📤 Exportar a Excel"

    # -------------------
    # Acción 2: Consulta SQL personalizada
    # -------------------
    def consulta_personalizada(self, request, queryset):
        # Obtener los CI de los objetos seleccionados
        ci_list = list(queryset.values_list("ci", flat=True))

        if not ci_list:
            self.message_user(request, "⚠️ Debes seleccionar al menos un operador.", level="warning")
            return

        # Construir placeholders (%s) dinámicamente según la cantidad de CI
        placeholders = ",".join(["%s"] * len(ci_list))

        sql = f"""
        SELECT ci, nombres, paterno, materno, celular, tipo_personal
        FROM usuarios_perfiloperador
        WHERE ci IN ({placeholders})
        ORDER BY ci;
        """

        with connection.cursor() as cursor:
            cursor.execute(sql, ci_list)  # pasar la lista como parámetros seguros
            columnas = [col[0] for col in cursor.description]  # nombres de columnas
            resultados = cursor.fetchall()

        # Crear Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Consulta SQL"

        # Encabezados
        ws.append([col.title() for col in columnas])

        # Filas
        for fila in resultados:
            ws.append([str(val) for val in fila])

        # Ajustar ancho
        for i, column_cells in enumerate(ws.columns, 1):
            length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[get_column_letter(i)].width = length + 2

        # Respuesta
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="consulta_personalizada.xlsx"'
        wb.save(response)
        return response

    consulta_personalizada.short_description = "🔍 Consulta Personalizada (SQL por CI)"


# -------------------
# Admin para Acta
# -------------------
@admin.register(Acta)
class ActaAdmin(admin.ModelAdmin):
    list_display = (
        'codigo', 'operador', 'provincia', 
        'municipio', 'localidad', 'recinto','actas_asignadas',
        'mesa1', 'mesa2', 'mesa3', 'mesa4', 
        'mesa5', 'mesa6', 'mesa7', 'mesa8'
    )
    search_fields = (
        'codigo', 'operador__ci', 'operador__nombres', 
        'operador__paterno', 'operador__materno'
    )
    list_filter = ('municipio',)

    # Mostrar operador con nombre completo
    def operador(self, obj):
        return f"{obj.operador.nombres} {obj.operador.paterno} {obj.operador.materno}"
    operador.short_description = "Operador"

# -------------------
# Admin para Actividad
# -------------------
@admin.register(Actividad)
class ActividadAdmin(admin.ModelAdmin):
    list_display = (
        'operador_ci',
        'operador_nombre',
        'get_coordinador', 'get_monitor', 'get_soporte', 
        'descripcion', 
        'estado', 
        'fecha', 
        'observacion',
        
    )
    search_fields = (
        'descripcion',
        'operador__ci',
        'operador__nombres',
        'operador__paterno',
        'operador__materno',
    )
    list_filter = (
    'estado',
    'descripcion',
    'operador__coordinador',
    'operador__monitor',
    'operador__soporte',
    )


    # -------------------
    # Coordinador / Soporte / Monitor (desde PerfilOperador)
    # -------------------
    def get_coordinador(self, obj):
        if obj.operador and obj.operador.coordinador:
            c = obj.operador.coordinador
            return f"{c.nombres} {c.paterno} {c.materno}"
        return "—"
    get_coordinador.short_description = "Coordinador"

    def get_soporte(self, obj):
        if obj.operador and obj.operador.soporte:
            s = obj.operador.soporte
            return f"{s.nombres} {s.paterno} {s.materno}"
        return "—"
    get_soporte.short_description = "Soporte"

    def get_monitor(self, obj):
        if obj.operador and obj.operador.monitor:
            m = obj.operador.monitor
            return f"{m.nombres} {m.paterno} {m.materno}"
        return "—"
    get_monitor.short_description = "Monitor"

    # -------------------
    # Operador
    # -------------------
    def operador_ci(self, obj):
        return obj.operador.ci if obj.operador else "-"
    operador_ci.short_description = "CI Operador"
    
    def operador_nombre(self, obj):
        if obj.operador:
            return f"{obj.operador.nombres} {obj.operador.paterno} {obj.operador.materno}"
        return "—"
    operador_nombre.short_description = "Nombre Operador"


# -------------------
# Admin para Observacion Adicional Operador
# -------------------
# -------------------
# Admin de ObservacionAdicional
# -------------------
@admin.register(ObservacionAdicional)
class ObservacionAdicionalAdmin(admin.ModelAdmin):
    list_display = (
        'operador_ci',
        'operador_completo',
        'observacion',
        'fecha',
    )
    search_fields = (
        'observacion',
        'operador__ci',
        'operador__nombres',
        'operador__paterno',
        'operador__materno',
    )
    list_filter = ('fecha',)

    # ✅ Habilita el buscador dinámico en el campo operador
    autocomplete_fields = ['operador']

    # -------------------
    # Campos personalizados en la tabla del admin
    # -------------------
    def operador_ci(self, obj):
        return obj.operador.ci
    operador_ci.short_description = "CI Operador"

    def operador_completo(self, obj):
        return f"{obj.operador.nombres} {obj.operador.paterno} {obj.operador.materno}"
    operador_completo.short_description = "Operador"



# -------------------
# Admin para DetalleAdicional
# -------------------
@admin.register(DetalleAdicional)
class DetalleAdicionalAdmin(admin.ModelAdmin):
    list_display = (
        'operador_ci',
        'operador_nombre',
        'usuario_apk',
        'transmitido',
        'no_transmitido',
    )
    search_fields = (
        'operador__ci',
        'operador__nombres',
        'operador__paterno',
        'operador__materno',
        'usuario_apk',
    )
    list_filter = ()

    # Mostrar CI del operador
    def operador_ci(self, obj):
        return obj.operador.ci
    operador_ci.short_description = "CI Operador"

    # Mostrar nombre completo del operador
    def operador_nombre(self, obj):
        return f"{obj.operador.nombres} {obj.operador.paterno} {obj.operador.materno}"
    operador_nombre.short_description = "Nombre Operador"

# -------------------
# Registrar Usuario
# -------------------
admin.site.register(Usuario, UsuarioAdmin)
